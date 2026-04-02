import json
import math
from pathlib import Path

import openpyxl
import requests


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "florida_customer_outage.xlsx"
ENERGY_WORKBOOK = ROOT / "florida_energy_consumption.xlsx"
NG_WORKBOOK = ROOT / "florida_ng_consumption.xlsx"
OUTPUT = ROOT / "florida_county_energy_map.html"
GEOJSON_URL = (
    "https://gis.fdacs.gov/hosting/rest/services/"
    "Florida_County_Boundaries/MapServer/0/query?where=1%3D1&outFields=*&f=geojson"
)


def dedupe_if_repeated(value):
    if value is None:
        return None
    if isinstance(value, float) and not value.is_integer():
        return value
    text = str(int(value)) if isinstance(value, (int, float)) else str(value)
    half = len(text) // 2
    if len(text) % 2 == 0 and text[:half] == text[half:]:
        return int(text[:half])
    return int(text)


def normalize_county_name(name):
    if not name:
        return name
    normalized = str(name).strip()
    aliases = {
        "DeSoto": "Desoto",
    }
    return aliases.get(normalized, normalized)


def normalize_workbook_data():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    sheet = workbook.active
    county_data = {}

    for county, tracked, customers_out, source_pct in sheet.iter_rows(
        min_row=2, values_only=True
    ):
        county = normalize_county_name(county)
        tracked_normalized = dedupe_if_repeated(tracked)
        out_raw = int(customers_out)
        out_deduped = dedupe_if_repeated(customers_out)

        raw_ratio = out_raw / tracked_normalized if tracked_normalized else 0
        deduped_ratio = out_deduped / tracked_normalized if tracked_normalized else 0
        out_normalized = (
            out_deduped
            if abs(deduped_ratio - source_pct) <= abs(raw_ratio - source_pct)
            else out_raw
        )

        county_data[county] = {
            "county": county,
            "customersTracked": tracked_normalized,
            "customersOut": out_normalized,
            "percentOutage": (
                out_normalized / tracked_normalized if tracked_normalized else 0
            ),
            "sourcePercentOutage": source_pct,
        }

    return county_data


def aggregate_yearly_consumption(workbook_path, source_name):
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    years = sorted({int(row[3]) for row in rows if row[3] is not None})
    values_by_year = {str(year): {} for year in years}

    for county, _state_id, _sector, year, _geo_id, source, consumption, _exp in rows:
        if year is None or source != source_name:
            continue
        normalized_county = normalize_county_name(county)
        year_key = str(int(year))
        values_by_year[year_key][normalized_county] = (
            values_by_year[year_key].get(normalized_county, 0) + float(consumption or 0)
        )

    return values_by_year, years[-1], years


def fetch_geometry():
    response = requests.get(GEOJSON_URL, timeout=30)
    response.raise_for_status()
    geojson = response.json()

    lat_ref = 27.8
    cos_lat_ref = math.cos(math.radians(lat_ref))

    def project(lon, lat):
        return [round(lon * cos_lat_ref, 5), round(-lat, 5)]

    features = []
    all_points = []
    for feature in geojson["features"]:
        name = feature["properties"].get("NAME")
        geometry = feature["geometry"]
        if not name or geometry["type"] not in {"Polygon", "MultiPolygon"}:
            continue

        source_polygons = (
            [geometry["coordinates"]]
            if geometry["type"] == "Polygon"
            else geometry["coordinates"]
        )

        polygons = []
        for polygon in source_polygons:
            rings = []
            for ring in polygon:
                projected_ring = [project(lon, lat) for lon, lat in ring]
                rings.append(projected_ring)
                all_points.extend(projected_ring)
            polygons.append(rings)

        features.append({"name": name.title(), "polygons": polygons})

    xs = [point[0] for point in all_points]
    ys = [point[1] for point in all_points]
    bounds = {
        "minX": round(min(xs), 5),
        "minY": round(min(ys), 5),
        "maxX": round(max(xs), 5),
        "maxY": round(max(ys), 5),
    }
    return features, bounds


def build_html(
    county_data,
    geometry,
    bounds,
    energy_year,
    energy_by_year,
    energy_years,
    natural_gas_year,
    natural_gas_by_year,
    natural_gas_years,
):
    data_json = json.dumps(county_data, separators=(",", ":"))
    geometry_json = json.dumps(geometry, separators=(",", ":"))
    bounds_json = json.dumps(bounds, separators=(",", ":"))
    energy_year_json = json.dumps(energy_by_year, separators=(",", ":"))
    energy_years_json = json.dumps(energy_years, separators=(",", ":"))
    natural_gas_year_json = json.dumps(natural_gas_by_year, separators=(",", ":"))
    natural_gas_years_json = json.dumps(natural_gas_years, separators=(",", ":"))

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
      <title>State of Florida Data Heatmaps</title>
  <style>
    :root {{
      --bg: #f4efe3;
      --panel: #0C2E3C;
      --panel-border: rgba(84, 71, 49, 0.16);
      --text: #C3E6F3;
      --muted: #C3E6F3;
      --accent: #1F6884;
      --accent-soft: #d9ece7;
      --map-water: rgba(160, 222, 235, 0.78);
      --county-stroke: rgba(28, 51, 60, 0.48);
      --shadow: 0 6px 18px #1F6884;
      --radius: 16px;
      --font: "Segoe UI", "Aptos", "Helvetica Neue", sans-serif;
    }}

    * {{
      box-sizing: border-box;
    }}

    body {{
      margin: 0;
      min-height: 100vh;
      font-family: var(--font);
      color: var(--text);
      background: #00151A;
    }}

    .app {{
      display: grid;
      grid-template-columns: 320px minmax(0, 1fr);
      grid-template-rows: minmax(0, 1fr) auto;
      min-height: calc(100vh - 48px);
      gap: 12px;
      padding: 24px;
    }}

    .panel {{
      background: var(--panel);
      border: 1px solid var(--panel-border);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      padding: 22px;
      backdrop-filter: blur(14px);
      display: flex;
      flex-direction: column;
      gap: 18px;
      font-size: 0.88em;
    }}

    .eyebrow {{
      font-size: 12px;
      font-weight: 700;
      letter-spacing: 0.12em;
      text-transform: uppercase;
      color: var(--accent);
    }}

    h1 {{
      margin: 0;
      font-size: 28px;
      line-height: 1.05;
      color: #1F6884;
    }}

    p {{
      margin: 0;
      color: var(--muted);
      line-height: 1.5;
    }}

    .section {{
      display: grid;
      gap: 10px;
    }}

    .section-title {{
      font-size: 12px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.12em;
      color: var(--accent);
    }}

    .layer-list {{
      display: grid;
      gap: 10px;
    }}

    .year-control {{
      display: grid;
      gap: 6px;
      margin-top: 2px;
    }}

    .year-control.hidden {{
      display: none;
    }}

    .year-control label {{
      font-size: 12px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.08em;
      color: var(--accent);
    }}

    .year-control select {{
      width: 100%;
      padding: 10px 12px;
      border-radius: 12px;
      border: 1px solid rgba(255, 255, 255, 0.14);
      background: rgba(255, 255, 255, 0.86);
      color: #0C2E3C;
      font: inherit;
      outline: none;
    }}

    label.option {{
      display: grid;
      grid-template-columns: auto 1fr;
      gap: 10px;
      align-items: start;
      padding: 12px 14px;
      border-radius: 16px;
      background: rgba(255, 255, 255, 0.72);
      border: 1px solid rgba(84, 71, 49, 0.08);
      cursor: pointer;
    }}

    label.option:hover {{
      border-color: rgba(0, 95, 115, 0.28);
      background: rgba(255, 255, 255, 0.92);
    }}

    .option-toggle {{
      margin-top: 2px;
      width: 18px;
      height: 18px;
      border: 2px solid rgba(31, 42, 44, 0.55);
      border-radius: 999px;
      background: #fff;
      display: inline-grid;
      place-content: center;
      cursor: pointer;
      padding: 0;
      flex: 0 0 auto;
    }}

    .option-toggle::before {{
      content: "";
      width: 8px;
      height: 8px;
      border-radius: 999px;
      transform: scale(0);
      transition: transform 120ms ease;
      background: #1f2a2c;
    }}

    .option-toggle.is-on::before {{
      transform: scale(1);
    }}

    .option-copy {{
      display: grid;
      gap: 4px;
    }}

    .option-copy strong {{
      font-size: 14px;
      color: var(--accent);
    }}

    .option-copy span {{
      color: var(--muted);
      font-size: 13px;
      line-height: 1.35;
    }}

    .stats {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 10px;
    }}

    .stat {{
      background: rgba(255, 255, 255, 0.72);
      border-radius: 16px;
      padding: 14px;
      border: 1px solid rgba(84, 71, 49, 0.08);
    }}

    .stat-label {{
      display: block;
      font-size: 12px;
      color: var(--accent);
      margin-bottom: 6px;
      text-transform: uppercase;
      letter-spacing: 0.08em;
    }}

    .stat-value {{
      font-size: 16.5px;
      font-weight: 700;
      color: var(--accent);
    }}

    .legend {{
      display: grid;
      gap: 10px;
    }}

    .legend-bar {{
      height: 14px;
      border-radius: 999px;
      background: linear-gradient(90deg, #FFF8DC 0%, #f1df9a 35%, #d0b246 68%, #A08000 100%);
      border: 1px solid rgba(84, 71, 49, 0.12);
    }}

    .legend-scale {{
      display: flex;
      justify-content: space-between;
      font-size: 12px;
      color: var(--muted);
      gap: 16px;
    }}

    .map-shell {{
      min-width: 0;
      position: relative;
      background: var(--map-water);
      border-radius: var(--radius);
      border: 1px solid rgba(84, 71, 49, 0.12);
      box-shadow: 0 10px 20px #1F6884;
      overflow: hidden;
      min-height: 0;
      height: calc(100vh - 48px);
    }}

    .map-topbar {{
      position: absolute;
      inset: 18px 18px auto auto;
      display: flex;
      align-items: center;
      gap: 10px;
      z-index: 4;
    }}

    .zoom-controls {{
      display: flex;
      gap: 8px;
    }}

    .zoom-controls button,
    .ghost-button {{
      width: 48px;
      height: 48px;
      border: 0;
      border-radius: 16px;
      background: var(--accent);
      color: #ffffff;
      font-size: 24px;
      cursor: pointer;
      box-shadow: 0 10px 24px rgba(17, 43, 52, 0.12);
    }}

    .ghost-button {{
      width: auto;
      padding: 0 16px;
      font-size: 14px;
      font-weight: 600;
    }}

    .zoom-controls button:hover,
    .ghost-button:hover {{
      background: #007991;
    }}

    svg {{
      display: block;
      width: 100%;
      height: 100%;
      min-height: 0;
      touch-action: none;
    }}

    .county {{
      stroke: var(--county-stroke);
      stroke-width: 0.045;
      vector-effect: non-scaling-stroke;
      transition: fill 180ms ease, opacity 180ms ease;
      cursor: pointer;
    }}

    .county:hover,
    .county.active {{
      stroke: rgba(20, 54, 66, 0.96);
      stroke-width: 0.085;
    }}

    .tooltip {{
      position: absolute;
      z-index: 6;
      pointer-events: none;
      min-width: 180px;
      max-width: 240px;
      padding: 12px 14px;
      border-radius: 14px;
      background: rgba(20, 54, 66, 0.94);
      color: #eff8f4;
      box-shadow: 0 16px 30px rgba(17, 43, 52, 0.24);
      transform: translate(-50%, calc(-100% - 16px));
      opacity: 0;
      transition: opacity 120ms ease;
    }}

    .tooltip.visible {{
      opacity: 1;
    }}

    .tooltip strong {{
      display: block;
      margin-bottom: 4px;
      font-size: 14px;
    }}

    .tooltip span {{
      display: block;
      font-size: 12px;
      line-height: 1.45;
      color: rgba(239, 248, 244, 0.82);
    }}

    .footer-panel {{
      grid-column: 1 / -1;
      background: var(--panel);
      border: 1px solid var(--panel-border);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      padding: 18px 22px;
    }}

    .footer-note {{
      font-size: 12px;
      color: var(--muted);
      line-height: 1.6;
    }}

    .footer-note ul {{
      margin: 8px 0 0;
      padding-left: 18px;
    }}

    .footer-note li + li {{
      margin-top: 6px;
    }}

    @media (max-width: 1080px) {{
      .app {{
        grid-template-columns: 1fr;
      }}

      .map-shell,
      svg {{
        height: 70vh;
      }}
    }}

    @media (max-width: 680px) {{
      .app {{
        padding: 14px;
        gap: 14px;
      }}

      .panel {{
        padding: 18px;
      }}

      .stats {{
        grid-template-columns: 1fr;
      }}
    }}
  </style>
</head>
<body>
  <div class="app">
    <aside class="panel">
      <div class="section">
        <div class="eyebrow">Single File</div>
        <h1>State of Florida Data Heatmaps</h1>
        <p>County-level data from a set of excel workbooks, embedded directly into this page with no external runtime dependencies.</p>
      </div>

      <div class="section">
        <div class="section-title">Data Layer</div>
        <div class="layer-list" id="layerList"></div>
        <div class="year-control hidden" id="layerYearControl">
          <label for="layerYearSelect" id="layerYearLabel">Layer Year</label>
          <select id="layerYearSelect" aria-label="Layer year"></select>
        </div>
      </div>

      <div class="section">
        <div class="section-title">Active Layer Summary</div>
        <div class="stats">
          <div class="stat">
            <span class="stat-label">Counties</span>
            <span class="stat-value" id="countyCount">0</span>
          </div>
          <div class="stat">
            <span class="stat-label">Non-Zero</span>
            <span class="stat-value" id="nonZeroCount">0</span>
          </div>
          <div class="stat">
            <span class="stat-label">Maximum</span>
            <span class="stat-value" id="maxValue">0</span>
          </div>
          <div class="stat">
            <span class="stat-label">Average</span>
            <span class="stat-value" id="avgValue">0</span>
          </div>
        </div>
      </div>

      <div class="section legend">
        <div class="section-title">Legend</div>
        <div class="legend-bar"></div>
        <div class="legend-scale">
          <span id="legendMin">Low</span>
          <span id="legendMax">High</span>
        </div>
      </div>

    </aside>

    <main class="map-shell">
      <div class="map-topbar">
        <div class="zoom-controls">
          <button type="button" id="zoomIn" aria-label="Zoom in">+</button>
          <button type="button" id="zoomOut" aria-label="Zoom out">-</button>
        </div>
        <button type="button" class="ghost-button" id="resetView">Reset view</button>
      </div>

      <div class="tooltip" id="tooltip"></div>

      <svg id="mapSvg" viewBox="0 0 100 100" role="img" aria-label="Florida county outage choropleth map">
        <g id="viewport">
          <g id="counties"></g>
        </g>
      </svg>
    </main>

    <footer class="footer-panel">
      <div class="footer-note">
        <div class="section-title">Data Sources</div>
        <ul>
          <li>County boundaries: Florida county boundary service published by the Florida Department of Agriculture and Consumer Services.</li>
          <li>2026 Outage data source: https://poweroutage.us/use-our-data/utilities</li>
          <li>{energy_years[0]} to {energy_years[-1]} Energy Consumption by County Data Source: https://maps.nlr.gov/slope/data-viewer</li>
          <li>{natural_gas_years[0]} to {natural_gas_years[-1]} Natural Gas Consumption by County Data Source: https://maps.nlr.gov/slope/data-viewer</li>
        </ul>
      </div>
    </footer>
  </div>

  <script>
    const COUNTY_DATA = {data_json};
    const COUNTY_GEOMETRY = {geometry_json};
    const MAP_BOUNDS = {bounds_json};
    const ENERGY_CONSUMPTION_BY_YEAR = {energy_year_json};
    const ENERGY_YEARS = {energy_years_json};
    const NATURAL_GAS_BY_YEAR = {natural_gas_year_json};
    const NATURAL_GAS_YEARS = {natural_gas_years_json};

    const LAYERS = [
      {{
        id: "percentOutage",
        label: "Electrical - Percent Outage",
        description: "",
        scaleMode: "quantileCurrent",
        format: (value) => `${{(value * 100).toFixed(value >= 0.01 ? 2 : 3)}}%`,
      }},
      {{
        id: "energyConsumptionMMBtu",
        label: "Electrical - Energy Consumption",
        description: "",
        scaleMode: "quantileGlobal",
        format: (value) => value.toLocaleString("en-US", {{ maximumFractionDigits: 0 }}) + " MMBtu",
      }},
      {{
        id: "naturalGasConsumptionMMBtu",
        label: "Natural Gas - Consumption",
        description: "",
        scaleMode: "quantileGlobal",
        format: (value) => value.toLocaleString("en-US", {{ maximumFractionDigits: 0 }}) + " MMBtu",
      }},
    ];

    const mapSvg = document.getElementById("mapSvg");
    const viewport = document.getElementById("viewport");
    const countiesGroup = document.getElementById("counties");
    const layerList = document.getElementById("layerList");
    const layerYearControl = document.getElementById("layerYearControl");
    const layerYearLabel = document.getElementById("layerYearLabel");
    const layerYearSelect = document.getElementById("layerYearSelect");
    const tooltip = document.getElementById("tooltip");
    const legendBar = document.querySelector(".legend-bar");

    const neutralFill = "#D6C1A3";
    const noDataFill = "#cfd8dc";
    const LAYER_STYLES = {{
      percentOutage: {{
        palette: ["#FFF8DC", "#f1df9a", "#dcbc63", "#c19a1f", "#A08000"],
        legend: "linear-gradient(90deg, #FFF8DC 0%, #f1df9a 35%, #d0b246 68%, #A08000 100%)",
      }},
      energyConsumptionMMBtu: {{
        palette: ["#F4B994", "#df915e", "#bb6430", "#a7511d", "#95420F"],
        legend: "linear-gradient(90deg, #F4B994 0%, #df915e 35%, #bb6430 68%, #95420F 100%)",
      }},
      naturalGasConsumptionMMBtu: {{
        palette: ["#d8efc6", "#9acc7a", "#58a64a", "#2f7d32", "#145a1f"],
        legend: "linear-gradient(90deg, #d8efc6 0%, #9acc7a 35%, #58a64a 68%, #145a1f 100%)",
      }},
    }};

    let heatmapEnabled = true;
    let activeLayer = "percentOutage";
    let selectedEnergyYear = {energy_year};
    let selectedNaturalGasYear = {natural_gas_year};
    let mapViewBox = null;
    const DEFAULT_SCALE = 1.44;
    const DEFAULT_OFFSET_X_PX = -150;
    const DEFAULT_OFFSET_Y_PX = 50;
    let currentTransform = {{ scale: DEFAULT_SCALE, x: 0, y: 0 }};
    let dragState = null;

    function computeViewBox() {{
      const padding = 1.5;
      const width = MAP_BOUNDS.maxX - MAP_BOUNDS.minX;
      const height = MAP_BOUNDS.maxY - MAP_BOUNDS.minY;
      return {{
        minX: MAP_BOUNDS.minX - padding,
        minY: MAP_BOUNDS.minY - padding,
        width: width + padding * 2,
        height: height + padding * 2,
      }};
    }}

    function buildPath(polygons) {{
      return polygons.map((polygon) => (
        polygon.map((ring) => ring.map(([x, y], index) => `${{index === 0 ? "M" : "L"}}${{x}} ${{y}}`).join(" ") + " Z").join(" ")
      )).join(" ");
    }}

    function polygonCentroid(ring) {{
      let area = 0;
      let x = 0;
      let y = 0;
      for (let i = 0; i < ring.length - 1; i += 1) {{
        const [x1, y1] = ring[i];
        const [x2, y2] = ring[i + 1];
        const cross = x1 * y2 - x2 * y1;
        area += cross;
        x += (x1 + x2) * cross;
        y += (y1 + y2) * cross;
      }}
      if (!area) {{
        const fallback = ring[Math.floor(ring.length / 2)] || [0, 0];
        return fallback;
      }}
      return [x / (3 * area), y / (3 * area)];
    }}

    function getCountyDatum(name) {{
      return COUNTY_DATA[name] || {{
        county: name,
        customersTracked: 0,
        customersOut: 0,
        percentOutage: 0,
        energyConsumptionMMBtu: 0,
        naturalGasConsumptionMMBtu: 0,
      }};
    }}

    function getLayerValue(countyName, layerId) {{
      if (layerId === "energyConsumptionMMBtu") {{
        return Number(ENERGY_CONSUMPTION_BY_YEAR[String(selectedEnergyYear)]?.[countyName]) || 0;
      }}
      if (layerId === "naturalGasConsumptionMMBtu") {{
        return Number(NATURAL_GAS_BY_YEAR[String(selectedNaturalGasYear)]?.[countyName]) || 0;
      }}
      return Number(getCountyDatum(countyName)[layerId]) || 0;
    }}

    function getLayerValues(layerId) {{
      return Object.keys(COUNTY_DATA)
        .map((countyName) => getLayerValue(countyName, layerId))
        .filter((value) => value > 0)
        .sort((a, b) => a - b);
    }}

    function getGlobalLayerValues(layerId) {{
      if (layerId === "energyConsumptionMMBtu") {{
        return Object.values(ENERGY_CONSUMPTION_BY_YEAR)
          .flatMap((yearValues) => Object.values(yearValues))
          .map((value) => Number(value) || 0)
          .filter((value) => value > 0)
          .sort((a, b) => a - b);
      }}
      if (layerId === "naturalGasConsumptionMMBtu") {{
        return Object.values(NATURAL_GAS_BY_YEAR)
          .flatMap((yearValues) => Object.values(yearValues))
          .map((value) => Number(value) || 0)
          .filter((value) => value > 0)
          .sort((a, b) => a - b);
      }}
      return getLayerValues(layerId);
    }}

    function quantile(sorted, q) {{
      if (!sorted.length) {{
        return 0;
      }}
      const position = (sorted.length - 1) * q;
      const base = Math.floor(position);
      const rest = position - base;
      const next = sorted[base + 1];
      return next === undefined
        ? sorted[base]
        : sorted[base] + rest * (next - sorted[base]);
    }}

    function buildScale(layerId) {{
      const layer = LAYERS.find((item) => item.id === layerId);
      const currentValues = getLayerValues(layerId);
      const scaleValues = layer?.scaleMode === "quantileGlobal"
        ? getGlobalLayerValues(layerId)
        : currentValues;

      if (!scaleValues.length) {{
        return {{
          thresholds: [],
          min: 0,
          max: 0,
          avg: 0,
          nonZeroCount: 0,
        }};
      }}

      return {{
        thresholds: [0.2, 0.4, 0.6, 0.8].map((q) => quantile(scaleValues, q)),
        min: currentValues[0] || 0,
        max: currentValues[currentValues.length - 1] || 0,
        avg: currentValues.length
          ? currentValues.reduce((sum, value) => sum + value, 0) / currentValues.length
          : 0,
        nonZeroCount: currentValues.length,
        legendMin: scaleValues[0] || 0,
        legendMax: scaleValues[scaleValues.length - 1] || 0,
      }};
    }}

    function getFillColor(value, scale) {{
      if (!Number.isFinite(value)) {{
        return noDataFill;
      }}
      if (value === 0) {{
        return neutralFill;
      }}
      const palette = LAYER_STYLES[activeLayer]?.palette || LAYER_STYLES.percentOutage.palette;
      const index = scale.thresholds.findIndex((threshold) => value <= threshold);
      return palette[index === -1 ? palette.length - 1 : index];
    }}

    function formatLayerValue(layerId, value) {{
      const layer = LAYERS.find((item) => item.id === layerId);
      return layer ? layer.format(value) : String(value);
    }}

    function getActiveLayerYears() {{
      if (activeLayer === "energyConsumptionMMBtu") {{
        return ENERGY_YEARS;
      }}
      if (activeLayer === "naturalGasConsumptionMMBtu") {{
        return NATURAL_GAS_YEARS;
      }}
      return [];
    }}

    function getSelectedYearForLayer(layerId) {{
      if (layerId === "energyConsumptionMMBtu") {{
        return selectedEnergyYear;
      }}
      if (layerId === "naturalGasConsumptionMMBtu") {{
        return selectedNaturalGasYear;
      }}
      return null;
    }}

    function renderLayerControls() {{
      layerList.innerHTML = "";
      for (const layer of LAYERS) {{
        const label = document.createElement("label");
        label.className = "option";
        label.innerHTML = `
          <button type="button" class="option-toggle ${{heatmapEnabled && activeLayer === layer.id ? "is-on" : ""}}" data-layer="${{layer.id}}" aria-pressed="${{heatmapEnabled && activeLayer === layer.id ? "true" : "false"}}" aria-label="${{layer.label}} toggle"></button>
          <div class="option-copy">
            <strong>${{layer.label}}</strong>
            ${{layer.description ? `<span>${{layer.description}}</span>` : ""}}
          </div>
        `;
        label.addEventListener("click", (event) => {{
          if (event.target.closest(".option-toggle") || event.target.closest(".option-copy")) {{
            if (activeLayer === layer.id) {{
              heatmapEnabled = !heatmapEnabled;
            }} else {{
              activeLayer = layer.id;
              heatmapEnabled = true;
            }}
          }}
          updateMap();
        }});
        layerList.appendChild(label);
      }}
    }}

    function renderEnergyYearControl() {{
      layerYearSelect.innerHTML = "";
      getActiveLayerYears().forEach((year) => {{
        const option = document.createElement("option");
        option.value = String(year);
        option.textContent = String(year);
        option.selected = Number(year) === Number(getSelectedYearForLayer(activeLayer));
        layerYearSelect.appendChild(option);
      }});
    }}

    function syncLayerControls() {{
      document.querySelectorAll(".option-toggle").forEach((toggle) => {{
        const isOn = heatmapEnabled && toggle.dataset.layer === activeLayer;
        toggle.classList.toggle("is-on", isOn);
        toggle.setAttribute("aria-pressed", isOn ? "true" : "false");
      }});
      const showLayerYear = activeLayer === "energyConsumptionMMBtu" || activeLayer === "naturalGasConsumptionMMBtu";
      layerYearControl.classList.toggle("hidden", !showLayerYear);
      if (showLayerYear) {{
        layerYearLabel.textContent = activeLayer === "energyConsumptionMMBtu" ? "Energy Year" : "Natural Gas Year";
        renderEnergyYearControl();
        layerYearSelect.value = String(getSelectedYearForLayer(activeLayer));
      }}
    }}

    function renderMap() {{
      countiesGroup.innerHTML = "";

      COUNTY_GEOMETRY.forEach((feature) => {{
        const countyDatum = getCountyDatum(feature.name);
        const path = document.createElementNS("http://www.w3.org/2000/svg", "path");
        path.setAttribute("class", "county");
        path.setAttribute("data-county", feature.name);
        path.setAttribute("d", buildPath(feature.polygons));
        path.addEventListener("mouseenter", (event) => showTooltip(event, countyDatum));
        path.addEventListener("mousemove", (event) => moveTooltip(event));
        path.addEventListener("mouseleave", hideTooltip);
        countiesGroup.appendChild(path);
      }});
    }}

    function updateLegend(scale) {{
      legendBar.style.background = LAYER_STYLES[activeLayer]?.legend || LAYER_STYLES.percentOutage.legend;
      document.getElementById("legendMin").textContent = heatmapEnabled
        ? formatLayerValue(activeLayer, scale.legendMin ?? scale.min)
        : "Heatmap off";
      document.getElementById("legendMax").textContent = heatmapEnabled
        ? formatLayerValue(activeLayer, scale.legendMax ?? scale.max)
        : "Neutral fill";
    }}

    function updateStats(scale) {{
      document.getElementById("countyCount").textContent = Object.keys(COUNTY_DATA).length.toLocaleString("en-US");
      document.getElementById("nonZeroCount").textContent = scale.nonZeroCount.toLocaleString("en-US");
      document.getElementById("maxValue").textContent = heatmapEnabled
        ? formatLayerValue(activeLayer, scale.max)
        : "Off";
      document.getElementById("avgValue").textContent = heatmapEnabled
        ? formatLayerValue(activeLayer, scale.avg)
        : "Off";
    }}

    function updateMap() {{
      const scale = buildScale(activeLayer);
      syncLayerControls();
      document.querySelectorAll(".county").forEach((path) => {{
        const countyName = path.getAttribute("data-county");
        const value = getLayerValue(countyName, activeLayer);
        path.setAttribute("fill", heatmapEnabled ? getFillColor(value, scale) : neutralFill);
      }});

      updateLegend(scale);
      updateStats(scale);
    }}

    function showTooltip(event, county) {{
      const energyValue = getLayerValue(county.county, "energyConsumptionMMBtu");
      const naturalGasValue = getLayerValue(county.county, "naturalGasConsumptionMMBtu");
      const activeLayerValue = getLayerValue(county.county, activeLayer);
      const activeLayerLabel = LAYERS.find((layer) => layer.id === activeLayer)?.label || activeLayer;
      let details = "";
      let heatmapLine = heatmapEnabled
        ? `<span>Heatmap: ${{activeLayerLabel}}: ${{formatLayerValue(activeLayer, activeLayerValue)}}</span>`
        : "<span>Heatmap: Off</span>";

      if (activeLayer === "percentOutage") {{
        details = `
          <span>Tracked: ${{county.customersTracked.toLocaleString("en-US")}}</span>
          <span>Customers Out: ${{county.customersOut.toLocaleString("en-US")}}</span>
          <span>Percent Outage: ${{formatLayerValue("percentOutage", county.percentOutage)}}</span>
        `;
      }} else if (activeLayer === "energyConsumptionMMBtu") {{
        details = `
          <span>Energy Consumption (${{selectedEnergyYear}}): ${{formatLayerValue("energyConsumptionMMBtu", energyValue)}}</span>
        `;
        heatmapLine = heatmapEnabled ? "" : "<span>Heatmap: Off</span>";
      }} else if (activeLayer === "naturalGasConsumptionMMBtu") {{
        details = `
          <span>Natural Gas Consumption (${{selectedNaturalGasYear}}): ${{formatLayerValue("naturalGasConsumptionMMBtu", naturalGasValue)}}</span>
        `;
        heatmapLine = heatmapEnabled ? "" : "<span>Heatmap: Off</span>";
      }}

      tooltip.innerHTML = `
        <strong>${{county.county}}</strong>
        ${{details}}
        ${{heatmapLine}}
      `;
      moveTooltip(event);
      tooltip.classList.add("visible");
    }}

    function moveTooltip(event) {{
      const shell = document.querySelector(".map-shell").getBoundingClientRect();
      tooltip.style.left = `${{event.clientX - shell.left}}px`;
      tooltip.style.top = `${{event.clientY - shell.top}}px`;
    }}

    function hideTooltip() {{
      tooltip.classList.remove("visible");
    }}

    function applyTransform() {{
      viewport.setAttribute(
        "transform",
        `translate(${{currentTransform.x}} ${{currentTransform.y}}) scale(${{currentTransform.scale}})`
      );
    }}

    function getMapCenter() {{
      if (!mapViewBox) {{
        return {{ x: 0, y: 0 }};
      }}
      return {{
        x: mapViewBox.minX + mapViewBox.width / 2,
        y: mapViewBox.minY + mapViewBox.height / 2,
      }};
    }}

    function getDefaultOffsetX() {{
      if (!mapViewBox) {{
        return 0;
      }}
      const rect = mapSvg.getBoundingClientRect();
      if (!rect.width) {{
        return 0;
      }}
      return (DEFAULT_OFFSET_X_PX / rect.width) * mapViewBox.width;
    }}

    function getDefaultOffsetY() {{
      if (!mapViewBox) {{
        return 0;
      }}
      const rect = mapSvg.getBoundingClientRect();
      if (!rect.height) {{
        return 0;
      }}
      return (DEFAULT_OFFSET_Y_PX / rect.height) * mapViewBox.height;
    }}

    function setZoom(nextScale, anchorX = null, anchorY = null) {{
      const clampedScale = Math.max(1, Math.min(12, nextScale));
      const center = getMapCenter();
      const resolvedAnchorX = anchorX ?? center.x;
      const resolvedAnchorY = anchorY ?? center.y;
      const scaleRatio = clampedScale / currentTransform.scale;
      currentTransform.x = resolvedAnchorX - (resolvedAnchorX - currentTransform.x) * scaleRatio;
      currentTransform.y = resolvedAnchorY - (resolvedAnchorY - currentTransform.y) * scaleRatio;
      currentTransform.scale = clampedScale;
      applyTransform();
    }}

    function resetView() {{
      currentTransform = {{ scale: DEFAULT_SCALE, x: 0, y: 0 }};
      const center = getMapCenter();
      currentTransform.x = center.x - center.x * currentTransform.scale + getDefaultOffsetX();
      currentTransform.y = center.y - center.y * currentTransform.scale + getDefaultOffsetY();
      applyTransform();
    }}

    function installPanZoom() {{
      mapSvg.addEventListener("wheel", (event) => {{
        event.preventDefault();
        const point = mapSvg.createSVGPoint();
        point.x = event.clientX;
        point.y = event.clientY;
        const svgPoint = point.matrixTransform(mapSvg.getScreenCTM().inverse());
        const delta = event.deltaY < 0 ? 1.12 : 0.89;
        setZoom(currentTransform.scale * delta, svgPoint.x, svgPoint.y);
      }}, {{ passive: false }});

      mapSvg.addEventListener("pointerdown", (event) => {{
        dragState = {{
          startX: event.clientX,
          startY: event.clientY,
          originX: currentTransform.x,
          originY: currentTransform.y,
        }};
        mapSvg.setPointerCapture(event.pointerId);
      }});

      mapSvg.addEventListener("pointermove", (event) => {{
        if (!dragState) {{
          return;
        }}
        const rect = mapSvg.getBoundingClientRect();
        const dx = (event.clientX - dragState.startX) / rect.width * 100;
        const dy = (event.clientY - dragState.startY) / rect.height * 100;
        currentTransform.x = dragState.originX + dx;
        currentTransform.y = dragState.originY + dy;
        applyTransform();
      }});

      function endDrag(event) {{
        if (!dragState) {{
          return;
        }}
        dragState = null;
        if (mapSvg.hasPointerCapture(event.pointerId)) {{
          mapSvg.releasePointerCapture(event.pointerId);
        }}
      }}

      mapSvg.addEventListener("pointerup", endDrag);
      mapSvg.addEventListener("pointercancel", endDrag);
    }}

    function initialize() {{
      const viewBox = computeViewBox();
      mapViewBox = viewBox;
      mapSvg.setAttribute("viewBox", `${{viewBox.minX}} ${{viewBox.minY}} ${{viewBox.width}} ${{viewBox.height}}`);

      renderLayerControls();
      renderEnergyYearControl();
      renderMap();
      resetView();
      updateMap();
      installPanZoom();

      layerYearSelect.addEventListener("change", (event) => {{
        if (activeLayer === "energyConsumptionMMBtu") {{
          selectedEnergyYear = Number(event.target.value);
        }} else if (activeLayer === "naturalGasConsumptionMMBtu") {{
          selectedNaturalGasYear = Number(event.target.value);
        }}
        updateMap();
      }});
      document.getElementById("zoomIn").addEventListener("click", () => setZoom(currentTransform.scale * 1.2));
      document.getElementById("zoomOut").addEventListener("click", () => setZoom(currentTransform.scale / 1.2));
      document.getElementById("resetView").addEventListener("click", resetView);
    }}

    initialize();
  </script>
</body>
</html>
"""


def main():
    county_data = normalize_workbook_data()
    energy_by_year, energy_year, energy_years = aggregate_yearly_consumption(
        ENERGY_WORKBOOK, "elec"
    )
    natural_gas_by_year, natural_gas_year, natural_gas_years = (
        aggregate_yearly_consumption(NG_WORKBOOK, "ng")
    )
    all_energy_counties = set()
    for year_values in energy_by_year.values():
        all_energy_counties.update(year_values.keys())
    all_natural_gas_counties = set()
    for year_values in natural_gas_by_year.values():
        all_natural_gas_counties.update(year_values.keys())
    for county in all_energy_counties | all_natural_gas_counties:
        county_data.setdefault(
            county,
            {
                "county": county,
                "customersTracked": 0,
                "customersOut": 0,
                "percentOutage": 0,
                "sourcePercentOutage": 0,
            },
        )
    for county_data_entry in county_data.values():
        county_data_entry["energyConsumptionMMBtu"] = (
            energy_by_year.get(str(energy_year), {}).get(county_data_entry["county"], 0)
        )
        county_data_entry["naturalGasConsumptionMMBtu"] = natural_gas_by_year.get(
            str(natural_gas_year), {}
        ).get(county_data_entry["county"], 0)
    geometry, bounds = fetch_geometry()
    html = build_html(
        county_data,
        geometry,
        bounds,
        energy_year,
        energy_by_year,
        energy_years,
        natural_gas_year,
        natural_gas_by_year,
        natural_gas_years,
    )
    OUTPUT.write_text(html, encoding="utf-8")
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
